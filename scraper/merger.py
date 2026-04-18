"""
Merge scraped_results.json back into the original Excel files.

Produces:
    KSA_Merchants_Riyadh_ENRICHED.xlsx
    KSA_Merchants_Jeddah_ENRICHED.xlsx
    ... etc.

Usage:
    python -m scraper.merger
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scraper.utils import normalize_url

# ── Paths ────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent.parent
RESULTS_JSON = BASE_DIR / "scraped_results.json"

EXCEL_FILES = [
    BASE_DIR / "KSA_Merchants_Riyadh.xlsx",
    BASE_DIR / "KSA_Merchants_Jeddah.xlsx",
    BASE_DIR / "KSA_Merchants_Dammam.xlsx",
    BASE_DIR / "KSA_Merchants_Khobar.xlsx",
    BASE_DIR / "KSA_Merchants_Mecca.xlsx",
    BASE_DIR / "KSA_Merchants_Medina.xlsx",
]

# New columns added to each sheet (in this order)
NEW_COLUMNS = [
    "has_hungerstation",
    "has_jahez",
    "has_talabat",
    "has_marsool",
    "instagram_url",
    "instagram_handle",
    "whatsapp_number",
    "email",
    "reservation_link",
    "founding_year",
    "branch_count_on_site",
    "has_loyalty_program",
    "has_mobile_app",
    "has_catering",
    "has_franchise_info",
    "has_online_menu",
    "delivery_platforms_count",
    "competitor_platforms",
    "is_website_alive",
    "load_time_ms",
    "page_language",
    "price_range_detected",
    "bd_enrichment_score",
]

DELIVERY_FLAGS = ["has_hungerstation", "has_jahez", "has_talabat", "has_marsool", "has_toyo", "has_noon_food"]
DELIVERY_LABELS = {
    "has_hungerstation": "HungerStation",
    "has_jahez": "Jahez",
    "has_talabat": "Talabat",
    "has_marsool": "Marsool",
    "has_toyo": "Toyo",
    "has_noon_food": "NoonFood",
}

# Header fill for new columns
_NEW_COL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_SCORE_FILL   = PatternFill(start_color="D6E4BC", end_color="D6E4BC", fill_type="solid")


# ── Score formula ────────────────────────────────────────────────────────────

def bd_score(r: dict[str, Any]) -> int:
    score = 0
    if r.get("is_alive") or r.get("is_website_alive"):
        score += 20
    if any(r.get(f) for f in ["has_hungerstation", "has_jahez", "has_talabat"]):
        score += 15
    if r.get("instagram_url"):
        score += 15
    if r.get("whatsapp_number"):
        score += 10
    if r.get("email"):
        score += 10
    if r.get("has_online_menu"):
        score += 10
    if r.get("founding_year"):
        score += 10
    if r.get("branch_count_on_site"):
        score += 10
    return score


def competitor_platforms(r: dict[str, Any]) -> str:
    return ",".join(
        DELIVERY_LABELS[f] for f in DELIVERY_FLAGS if r.get(f)
    )


def delivery_count(r: dict[str, Any]) -> int:
    return sum(1 for f in DELIVERY_FLAGS if r.get(f))


# ── Build lookup ──────────────────────────────────────────────────────────────

def build_lookup(results_path: Path) -> dict[str, dict[str, Any]]:
    """Returns {normalized_url: enriched_dict}."""
    if not results_path.exists():
        print(f"[ERROR] {results_path} not found. Run batch_runner first.")
        sys.exit(1)

    with open(results_path, encoding="utf-8") as f:
        records = json.load(f)

    lookup: dict[str, dict[str, Any]] = {}
    for rec in records:
        url = rec.get("website_url") or rec.get("original_url", "")
        if not url:
            continue
        try:
            key = normalize_url(url)
        except Exception:
            key = url
        # Flatten + augment
        rec["is_website_alive"] = rec.get("is_alive", False)
        rec["delivery_platforms_count"] = delivery_count(rec)
        rec["competitor_platforms"] = competitor_platforms(rec)
        rec["bd_enrichment_score"] = bd_score(rec)
        lookup[key] = rec

    return lookup


# ── Excel enrichment ──────────────────────────────────────────────────────────

def enrich_file(src: Path, lookup: dict[str, dict[str, Any]]) -> None:
    dest = src.with_name(src.stem + "_ENRICHED.xlsx")
    print(f"  Processing {src.name} → {dest.name}")

    # Load with openpyxl to preserve formatting
    wb = openpyxl.load_workbook(src)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the Website column index (1-based)
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        web_col_idx = next(
            (i + 1 for i, h in enumerate(header_row) if h and "website" in str(h).lower()),
            None,
        )
        if web_col_idx is None:
            continue  # sheet has no Website column

        # Find the first empty column after existing data
        last_col = ws.max_column
        start_new_col = last_col + 1

        # Write new headers
        for j, col_name in enumerate(NEW_COLUMNS):
            cell = ws.cell(row=1, column=start_new_col + j, value=col_name)
            if col_name == "bd_enrichment_score":
                cell.fill = _SCORE_FILL
            else:
                cell.fill = _NEW_COL_FILL

        # Write data rows
        for row_idx in range(2, ws.max_row + 1):
            raw_url = ws.cell(row=row_idx, column=web_col_idx).value
            if not raw_url:
                continue
            try:
                key = normalize_url(str(raw_url).strip())
            except Exception:
                continue

            rec = lookup.get(key)
            if not rec:
                continue

            for j, col_name in enumerate(NEW_COLUMNS):
                val = rec.get(col_name)
                # Booleans → YES/NO for readability
                if isinstance(val, bool):
                    val = "YES" if val else "NO"
                ws.cell(row=row_idx, column=start_new_col + j, value=val)

        # Auto-size new columns (cap at 50)
        for j, col_name in enumerate(NEW_COLUMNS):
            col_letter = get_column_letter(start_new_col + j)
            ws.column_dimensions[col_letter].width = min(
                max(len(col_name) + 2, 12), 50
            )

    wb.save(dest)
    print(f"    Saved {dest.name}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Building lookup from scraped_results.json…")
    lookup = build_lookup(RESULTS_JSON)
    print(f"  Lookup entries: {len(lookup):,}")

    matched = 0
    for src in EXCEL_FILES:
        if not src.exists():
            print(f"[WARN] Missing: {src.name}")
            continue
        enrich_file(src, lookup)
        matched += 1

    print(f"\n✓ Enriched {matched} Excel files.")


if __name__ == "__main__":
    main()
