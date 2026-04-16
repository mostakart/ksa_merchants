# ksa_pipeline_master.py
# ═══════════════════════════════════════════════════════════════════════════════
# KSA MERCHANT INTELLIGENCE PIPELINE — ALL-IN-ONE
# ═══════════════════════════════════════════════════════════════════════════════
#
# Usage:
#   python ksa_pipeline_master.py                    # Run all cities
#   python ksa_pipeline_master.py --city Jeddah      # Run one city only
#   python ksa_pipeline_master.py --fresh            # Start from scratch
#   python ksa_pipeline_master.py --analyze-only     # Skip pipeline, just analyze
#   python ksa_pipeline_master.py --export-only      # Re-export Excel from checkpoint
#
# Phases (automatic, per city):
#   1. Discovery   — Find malls + merchants (Google Places Text Search)
#   2. Enrichment  — Rating, reviews, price, reviews text (Place Details Basic)
#   3. Contact     — Phone, website, opening hours, exact lat/lng (Place Details Contact)
#   4. Export      — Excel with priority scoring + color coding
#   5. Analysis    — BD scores, sentiment, whitespace gaps, CSV reports
#
# Checkpoint: saves every 10 merchants — safe to Ctrl+C anytime
# Cost: ~$20 per city (Riyadh = 1385 merchants @ ~$0.020/merchant)

import os, sys, re, json, math, time, logging, warnings
import requests
from pathlib import Path
from collections import Counter
from dotenv import load_dotenv

warnings.filterwarnings("ignore")
load_dotenv()

# ── CONFIG ─────────────────────────────────────────────────────────────────────
GOOGLE_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
BASE_URL       = "https://maps.googleapis.com/maps/api/place"
OSM_URL        = "https://overpass-api.de/api/interpreter"
SAVE_EVERY     = 10          # Save checkpoint every N merchants
REQUEST_GAP    = 1.5         # Seconds between Google API calls
OUTPUT_DIR     = Path("output")

KSA_CITIES = [
    {"name": "Riyadh",  "lat": 24.7136, "lng": 46.6753},
    {"name": "Jeddah",  "lat": 21.4858, "lng": 39.1925},
    {"name": "Dammam",  "lat": 26.4207, "lng": 50.0888},
    {"name": "Khobar",  "lat": 26.2794, "lng": 50.2083},
    {"name": "Mecca",   "lat": 21.3891, "lng": 39.8579},
    {"name": "Medina",  "lat": 24.5247, "lng": 39.5692},
]

FB_CATEGORIES = [
    ("restaurant", "Casual Dining"),
    ("cafe",       "Cafes"),
    ("fast food",  "Fast Food"),
    ("bakery",     "Cafes"),
    ("dessert",    "Desserts & Sweets"),
    ("مطعم",       "Casual Dining"),
    ("كافيه",      "Cafes"),
    ("حلويات",     "Desserts & Sweets"),
]

COST_PER = {
    "text_search":   0.032,
    "place_details": 0.017,
    "contact":       0.003,
    "gemini":        0.001,
}

# ── LOGGING ───────────────────────────────────────────────────────────────────
OUTPUT_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / "pipeline.log"),
    ]
)
logger = logging.getLogger(__name__)

# ── COST TRACKER ──────────────────────────────────────────────────────────────
_costs = {k: 0 for k in COST_PER}

def track(call_type: str, n: int = 1):
    _costs[call_type] = _costs.get(call_type, 0) + n

def cost_usd() -> float:
    return sum(_costs[k] * COST_PER[k] for k in COST_PER)

def cost_line() -> str:
    total = cost_usd()
    parts = [f"{k}: {_costs[k]} (${_costs[k]*COST_PER[k]:.3f})" for k in COST_PER if _costs[k]>0]
    return f"💰 ${total:.3f}  [{' | '.join(parts)}]"


# ══════════════════════════════════════════════════════════════════════════════
# CHECKPOINT
# ══════════════════════════════════════════════════════════════════════════════
CHECKPOINT_FILE = OUTPUT_DIR / "master_checkpoint.json"

def load_checkpoint() -> dict:
    if CHECKPOINT_FILE.exists():
        try:
            data = json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
            cities_done = data.get("cities_done", [])
            total = sum(len(v) for v in data.get("enriched", {}).values())
            logger.info(f"✅ Checkpoint: {total} merchants done | cities: {cities_done}")
            return data
        except Exception as e:
            logger.warning(f"Checkpoint corrupted, starting fresh: {e}")
    return {
        "cities_done":    [],
        "mall_data":      {},
        "enriched":       {},
        "contact_done":   [],
        "costs_snapshot": {},
    }

def save_checkpoint(state: dict):
    state["costs_snapshot"] = dict(_costs)
    CHECKPOINT_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def clear_checkpoint():
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()
    logger.info("🗑️  Checkpoint cleared")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════
def discover_malls(city: dict) -> list[dict]:
    if not GOOGLE_API_KEY: return []
    params = {
        "query":        f"shopping mall في {city['name']} السعودية",
        "key":          GOOGLE_API_KEY,
        "language":     "ar",
        "region":       "sa",
        "locationbias": f"circle:30000@{city['lat']},{city['lng']}",
    }
    try:
        resp = requests.get(f"{BASE_URL}/textsearch/json", params=params, timeout=12)
        resp.raise_for_status()
        track("text_search")
        malls = []
        for r in resp.json().get("results", []):
            malls.append({
                "mall_name": r.get("name"),
                "city":      city["name"],
                "place_id":  r.get("place_id"),
                "lat":       r["geometry"]["location"]["lat"],
                "lng":       r["geometry"]["location"]["lng"],
            })
        time.sleep(REQUEST_GAP)
        return malls
    except Exception as e:
        logger.error(f"Mall discovery failed [{city['name']}]: {e}")
        return []


def discover_merchants(mall: dict) -> list[dict]:
    if not GOOGLE_API_KEY: return []
    merchants, seen = [], set()

    for cat_query, cat_label in FB_CATEGORIES:
        params = {
            "query":        f"{cat_query} في {mall['mall_name']} {mall['city']}",
            "key":          GOOGLE_API_KEY,
            "language":     "ar",
            "region":       "sa",
            "locationbias": f"circle:500@{mall['lat']},{mall['lng']}",
        }
        try:
            resp = requests.get(f"{BASE_URL}/textsearch/json", params=params, timeout=12)
            resp.raise_for_status()
            track("text_search")
            for r in resp.json().get("results", []):
                pid = r.get("place_id")
                if pid in seen: continue
                seen.add(pid)
                merchants.append({
                    "merchant":      r.get("name"),
                    "mall_name":     mall["mall_name"],
                    "city":          mall["city"],
                    "mall_lat":      mall["lat"],
                    "mall_lng":      mall["lng"],
                    "place_id":      pid,
                    "rating":        r.get("rating"),
                    "total_reviews": r.get("user_ratings_total"),
                    "category":      cat_label,
                    "address":       r.get("formatted_address", ""),
                })
            time.sleep(REQUEST_GAP)
        except Exception as e:
            logger.error(f"Merchant discovery failed [{cat_query}@{mall['mall_name']}]: {e}")

    return merchants


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — ENRICHMENT (rating, reviews, price, reviews text)
# ══════════════════════════════════════════════════════════════════════════════
_details_cache = {}

def get_place_details(place_id: str) -> dict:
    if not place_id or not GOOGLE_API_KEY: return {}
    if place_id in _details_cache: return _details_cache[place_id]

    params = {
        "place_id": place_id,
        "key":      GOOGLE_API_KEY,
        "language": "ar",
        "fields":   "name,rating,user_ratings_total,price_level,reviews,geometry",
    }
    try:
        resp = requests.get(f"{BASE_URL}/details/json", params=params, timeout=12)
        data = resp.json()
        if data.get("status") == "OK":
            result = data.get("result", {})
            _details_cache[place_id] = result
            track("place_details")
            time.sleep(REQUEST_GAP)
            return result
    except Exception as e:
        logger.warning(f"Place details failed [{place_id}]: {e}")
    return {}


def clean_review(text: str) -> str:
    if not text: return ""
    text = re.sub(r"\s+", " ", text).strip()
    return text[:120] + "…" if len(text) > 120 else text

def price_to_range(level) -> str:
    return {0:"10–20 SAR",1:"20–40 SAR",2:"40–80 SAR",3:"80–150 SAR",4:"150+ SAR"}.get(level,"")

def enrich_basic(merchant: dict) -> dict:
    details = get_place_details(merchant.get("place_id",""))
    geo     = details.get("geometry",{}).get("location",{})
    raw_rev = details.get("reviews",[])
    reviews_text = " | ".join(
        clean_review(r.get("text","")) for r in raw_rev[:3] if r.get("text")
    )
    pl = details.get("price_level")
    return {
        **merchant,
        "rating":        details.get("rating")              or merchant.get("rating"),
        "total_reviews": details.get("user_ratings_total")  or merchant.get("total_reviews"),
        "price_level":   pl,
        "avg_price":     price_to_range(pl),
        "lat":           geo.get("lat") or merchant.get("mall_lat"),
        "lng":           geo.get("lng") or merchant.get("mall_lng"),
        "reviews_text":  reviews_text,
        "p1_status":     "OK" if details else "PARTIAL",
    }


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — CONTACT ENRICHMENT (phone, website, opening hours)
# ══════════════════════════════════════════════════════════════════════════════
def get_contact_details(place_id: str) -> dict:
    if not place_id or not GOOGLE_API_KEY: return {}

    params = {
        "place_id": place_id,
        "key":      GOOGLE_API_KEY,
        "language": "ar",
        "fields":   "formatted_phone_number,website,opening_hours,geometry,price_level",
    }
    try:
        resp = requests.get(f"{BASE_URL}/details/json", params=params, timeout=12)
        data = resp.json()
        if data.get("status") == "OK":
            r     = data.get("result", {})
            geo   = r.get("geometry",{}).get("location",{})
            hours = r.get("opening_hours",{})
            track("contact")
            time.sleep(REQUEST_GAP)
            return {
                "phone":         r.get("formatted_phone_number",""),
                "website":       r.get("website",""),
                "opening_hours": ", ".join(hours.get("weekday_text",[])),
                "lat":           geo.get("lat"),
                "lng":           geo.get("lng"),
                "price_level":   r.get("price_level"),
            }
    except Exception as e:
        logger.warning(f"Contact failed [{place_id}]: {e}")
    return {}


def enrich_contact(merchant: dict) -> dict:
    contact = get_contact_details(merchant.get("place_id",""))
    m = dict(merchant)
    m["phone"]         = contact.get("phone","")
    m["website"]       = contact.get("website","")
    m["opening_hours"] = contact.get("opening_hours","")
    if contact.get("lat"):  m["lat"] = contact["lat"]
    if contact.get("lng"):  m["lng"] = contact["lng"]
    if not m.get("avg_price") and contact.get("price_level") is not None:
        m["price_level"] = contact["price_level"]
        m["avg_price"]   = price_to_range(contact["price_level"])
    m["p2_status"] = "OK" if contact else "PARTIAL"
    return m


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY SCORING (inline, no pandas needed)
# ══════════════════════════════════════════════════════════════════════════════
def calc_priority(merchant: dict, branch_count: int) -> str:
    score   = 0
    rating  = float(merchant.get("rating") or 0)
    reviews = int(merchant.get("total_reviews") or 0)
    price   = int(merchant.get("price_level") or 0)

    if rating  >= 4.5:   score += 30
    elif rating >= 4.0:  score += 20
    elif rating >= 3.5:  score += 10

    if reviews >= 1000:  score += 30
    elif reviews >= 500: score += 20
    elif reviews >= 100: score += 10

    if branch_count >= 8:   score += 25
    elif branch_count >= 4: score += 15
    elif branch_count >= 2: score += 8

    if price >= 3:   score += 15
    elif price >= 2: score += 10
    elif price >= 1: score += 5

    if score >= 60: return "🔴 High"
    if score >= 35: return "🟡 Medium"
    return "🟢 Low"


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(all_enriched: dict, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    PRIORITY_COLORS  = {"🔴 High":"FFFF4444","🟡 Medium":"FFFFFF88","🟢 Low":"FF88FF88"}
    CATEGORY_COLORS  = {"Casual Dining":"FF4A86E8","Fast Food":"FF6AA84F","Cafes":"FF9FC5E8","Desserts & Sweets":"FFFF9900","Other":"FFFFFFFF"}

    COLUMNS = [
        ("Priority",12),("Category",14),("Merchant",28),("Mall",22),("City",10),
        ("Rating",8),("Reviews",10),("Avg Price",14),("Branches (KSA)",14),
        ("Phone",18),("Website",30),("Opening Hours",45),("Lat",12),("Lng",12),
        ("Top Reviews",70),
    ]

    # Build branch count map across ALL merchants
    all_flat = [m for merchants in all_enriched.values() for m in merchants]
    branch_map = Counter(m.get("merchant","") for m in all_flat)

    wb = Workbook()
    wb.remove(wb.active)

    # One sheet per city (combine malls)
    city_batches = {}
    for mall_name, merchants in all_enriched.items():
        for m in merchants:
            city = m.get("city","Unknown")
            city_batches.setdefault(city, {}).setdefault(mall_name,[]).append(m)

    for city, malls in city_batches.items():
        for mall_name, merchants in malls.items():
            safe_title = re.sub(r'[:\\/?*\[\]]', '', mall_name)[:31]
            ws = wb.create_sheet(title=safe_title)
            for i,(col,w) in enumerate(COLUMNS,1):
                cell = ws.cell(row=1,column=i,value=col)
                cell.font=Font(name="Arial",bold=True,size=10)
                cell.border=Border(bottom=Side(style="thin"))
                cell.alignment=Alignment(horizontal="center",vertical="center")
                ws.column_dimensions[get_column_letter(i)].width=w
            ws.freeze_panes="A2"

            for rn,m in enumerate(merchants,start=2):
                name  = m.get("merchant","")
                bc    = branch_map.get(name,1)
                pri   = calc_priority(m, bc)
                cat   = m.get("category","Other")
                rev_t = m.get("reviews_text","")

                vals = [
                    pri, cat, name,
                    m.get("mall_name",""), m.get("city",""),
                    m.get("rating",""), m.get("total_reviews",""),
                    m.get("avg_price",""), bc,
                    m.get("phone",""), m.get("website",""),
                    m.get("opening_hours",""),
                    m.get("lat",""), m.get("lng",""),
                    rev_t,
                ]
                for ci,v in enumerate(vals,1):
                    cell=ws.cell(row=rn,column=ci,value=v)
                    cell.font=Font(name="Arial",size=10)
                    if ci==1: cell.fill=PatternFill("solid",fgColor=PRIORITY_COLORS.get(pri,"FFFFFFFF"))
                    elif ci==2: cell.fill=PatternFill("solid",fgColor=CATEGORY_COLORS.get(cat,"FFFFFFFF"))
                    cell.alignment=Alignment(horizontal="left" if ci==15 else "center",vertical="center",wrap_text=(ci==15))
                ws.row_dimensions[rn].height=22.5

    wb.save(output_path)
    logger.info(f"✅ Excel saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 5 — ANALYSIS (embedded from analyze_ksa_merchants_v2)
# ══════════════════════════════════════════════════════════════════════════════
POSITIVE_KEYWORDS = ["لذيذ","لذيذة","رائع","رائعة","ممتاز","ممتازة","جميل","جميلة","نظيف","نظيفة","سريع","سريعة","مذهل","إبداع","طازج","طازجة","هادئ","مريح","أنصح","ننصح","أفضل","أحسن","جودة","يستاهل","احترافي","متميز","خدمة ممتازة","تجربة رائعة","تنوع","كرم","شهي","مبدع"]
NEGATIVE_KEYWORDS = ["بطيء","بطيئة","انتظار","غالي","غالية","سيء","سيئة","مزعج","بارد","باردة","ضوضاء","صغير","ضيق","مشكلة","شكوى","خيبة","متأخر","قديم","قذر","مكلف","مخيب","لا أنصح","لا ننصح","مؤسف","أسعار مرتفعة","جودة رديئة","لن أعود","ما أرجع","ضعيف","بدون طعم","رديء","فوضى","تأخر"]
CATEGORY_PRICE_DEFAULTS = {"Casual Dining":65.0,"Fast Food":35.0,"Cafes":45.0,"Desserts & Sweets":30.0,"Other":50.0}

def run_analysis(excel_path: str):
    try:
        import pandas as pd
        import numpy as np
    except ImportError:
        logger.error("pandas not installed — run: pip install pandas openpyxl")
        return

    logger.info("\n📊 Running analysis...")
    xl = pd.ExcelFile(excel_path)
    frames = [xl.parse(s) for s in xl.sheet_names]
    df = pd.concat(frames, ignore_index=True)

    # Rename
    rename = {"Merchant":"merchant","Mall":"mall","City":"city","Category":"category",
              "Rating":"rating","Reviews":"reviews","Avg Price":"avg_price",
              "Branches (KSA)":"branches","Phone":"phone","Website":"website",
              "Opening Hours":"opening_hours","Lat":"lat","Lng":"lng",
              "Priority":"priority","Top Reviews":"reviews_text"}
    df = df.rename(columns={k:v for k,v in rename.items() if k in df.columns})

    df["rating"]   = pd.to_numeric(df.get("rating"),  errors="coerce")
    df["reviews"]  = pd.to_numeric(df.get("reviews"), errors="coerce").fillna(0).astype(int)
    df["branches"] = pd.to_numeric(df.get("branches"),errors="coerce").fillna(1).astype(int)
    df["lat"]      = pd.to_numeric(df.get("lat"),     errors="coerce")
    df["lng"]      = pd.to_numeric(df.get("lng"),     errors="coerce")

    # Priority clean
    def clean_pri(p):
        p=str(p)
        if "High" in p or "🔴" in p: return "High"
        if "Low"  in p or "🟢" in p: return "Low"
        return "Medium"
    df["priority"] = df.get("priority",pd.Series(["Medium"]*len(df))).apply(clean_pri)

    # Price
    def parse_price(p):
        if pd.isna(p) or str(p).strip() in ["","nan"]: return None
        nums=re.findall(r"\d+",str(p))
        if len(nums)>=2: return (int(nums[0])+int(nums[1]))/2
        if len(nums)==1: return int(nums[0])
        return None
    def price_seg(mid):
        if pd.isna(mid): return "Unknown"
        if mid<40: return "Budget"
        if mid<80: return "Mid"
        if mid<150: return "Premium"
        return "Luxury"
    df["price_mid_raw"] = df.get("avg_price",pd.Series([None]*len(df))).apply(parse_price)
    cat_medians = df.groupby("category")["price_mid_raw"].median().to_dict()
    df["price_mid"] = df.apply(lambda r: r["price_mid_raw"] if pd.notna(r["price_mid_raw"]) else cat_medians.get(r.get("category"),CATEGORY_PRICE_DEFAULTS.get(r.get("category"),50)), axis=1)
    df["price_segment"] = df["price_mid"].apply(price_seg)

    # Sentiment
    def sentiment(text):
        t=str(text) if pd.notna(text) else ""
        pos=[kw for kw in POSITIVE_KEYWORDS if kw in t]
        neg=[kw for kw in NEGATIVE_KEYWORDS if kw in t]
        total=len(pos)+len(neg)
        ratio=len(pos)/total if total>0 else 0.5
        label = "Positive" if ratio>=0.7 else ("Mixed" if ratio>=0.4 else "Negative")
        return {"pos":len(pos),"neg":len(neg),"ratio":round(ratio,2),"label":label,
                "top_pos":", ".join(pos[:3]),"top_neg":", ".join(neg[:3]),
                "delivery":any(w in t for w in ["توصيل","ديليفري","delivery"])}
    sent = df.get("reviews_text",pd.Series([""]*len(df))).apply(sentiment)
    df["pos_kw_count"]     = sent.apply(lambda x: x["pos"])
    df["neg_kw_count"]     = sent.apply(lambda x: x["neg"])
    df["sentiment_ratio"]  = sent.apply(lambda x: x["ratio"])
    df["sentiment_label"]  = sent.apply(lambda x: x["label"])
    df["top_pos_kw"]       = sent.apply(lambda x: x["top_pos"])
    df["top_neg_kw"]       = sent.apply(lambda x: x["top_neg"])
    df["mentions_delivery"]= sent.apply(lambda x: x["delivery"])

    # Geo zone
    def geo_zone(row):
        lat,lng=row.get("lat"),row.get("lng")
        if pd.isna(lat) or pd.isna(lng): return "Unknown"
        if lat>24.75: return "North"
        if lat<24.60: return "South"
        if lng>46.75: return "East"
        if lng<46.55: return "West"
        return "Central"
    df["geo_zone"] = df.apply(geo_zone, axis=1)

    # BD Score
    max_rev = df["reviews"].max() or 1
    def bd_score(row):
        r=float(row.get("rating") or 0)
        v=int(row.get("reviews") or 0)
        sr=float(row.get("sentiment_ratio") or 0.5)
        b=min(int(row.get("branches") or 1),13)
        p={"High":1.0,"Medium":0.5,"Low":0.0}.get(row.get("priority","Medium"),0.5)
        has_web=1 if str(row.get("website","")) not in ["","nan"] else 0
        score = (r/5)*25 + (math.log10(v+1)/math.log10(max_rev+1))*20 + sr*15 + (b/13)*15 + p*10 + has_web*5
        return round(min(score*1.05,100) if row.get("mentions_delivery") else min(score,100),2)
    df["bd_priority_score"] = df.apply(bd_score, axis=1)
    df["tier"] = df["bd_priority_score"].apply(lambda s: "Tier 1" if s>=75 else ("Tier 2" if s>=50 else "Tier 3"))

    med_r = df["rating"].median()
    med_v = df["reviews"].median()
    def quadrant(row):
        r,v=row["rating"],row["reviews"]
        if pd.isna(r): return "Unknown"
        if r>=med_r and v>=med_v: return "⭐ Star"
        if r>=med_r and v<med_v:  return "💎 Hidden Gem"
        if r<med_r  and v>=med_v: return "⚠️ Risky"
        return "❌ Weak"
    df["quadrant"] = df.apply(quadrant,axis=1)

    # ── Save CSVs ──────────────────────────────────────────────────────────────
    analysis_dir = OUTPUT_DIR / "analysis"
    analysis_dir.mkdir(exist_ok=True)

    df.to_csv(analysis_dir/"merchants_enriched.csv", index=False, encoding="utf-8-sig")

    # Mall metrics
    mall_agg = df.groupby("mall").agg(
        merchant_count=("merchant","count"),
        avg_rating    =("rating","mean"),
        avg_reviews   =("reviews","mean"),
        total_reviews =("reviews","sum"),
        avg_bd_score  =("bd_priority_score","mean"),
        pct_tier1     =("tier",lambda x:(x=="Tier 1").mean()*100),
        pct_high_pri  =("priority",lambda x:(x=="High").mean()*100),
        pct_positive  =("sentiment_label",lambda x:(x=="Positive").mean()*100),
        pct_delivery  =("mentions_delivery","mean"),
        whitespace_gap=("category","nunique"),
    ).round(2).reset_index()
    mall_agg.to_csv(analysis_dir/"metrics_by_mall.csv", index=False, encoding="utf-8-sig")

    # Category metrics
    cat_agg = df.groupby("category").agg(
        merchant_count=("merchant","count"),
        avg_rating    =("rating","mean"),
        avg_bd_score  =("bd_priority_score","mean"),
        avg_branches  =("branches","mean"),
        pct_tier1     =("tier",lambda x:(x=="Tier 1").mean()*100),
        pct_positive  =("sentiment_label",lambda x:(x=="Positive").mean()*100),
    ).round(2).reset_index()
    cat_agg.to_csv(analysis_dir/"metrics_by_category.csv", index=False, encoding="utf-8-sig")

    # Whitespace gaps
    existing = set(zip(df["mall"],df["category"]))
    mall_scores = df.groupby("mall")["bd_priority_score"].mean().to_dict()
    gaps = []
    for mall in df["mall"].unique():
        for cat in df["category"].unique():
            if (mall,cat) not in existing:
                s = mall_scores.get(mall,0)
                gaps.append({"mall":mall,"missing_category":cat,"opportunity_score":round(s,2)})
    pd.DataFrame(gaps).sort_values("opportunity_score",ascending=False).to_csv(
        analysis_dir/"whitespace_gaps.csv", index=False, encoding="utf-8-sig")

    # Keywords corpus
    pos_all,neg_all=[],[]
    for _,row in df.iterrows():
        pos_all.extend([w.strip() for w in str(row.get("top_pos_kw","")).split(",") if w.strip()])
        neg_all.extend([w.strip() for w in str(row.get("top_neg_kw","")).split(",") if w.strip()])
    kw_rows=[]
    for kw,f in Counter(pos_all).most_common(30): kw_rows.append({"keyword":kw,"frequency":f,"sentiment":"positive"})
    for kw,f in Counter(neg_all).most_common(30): kw_rows.append({"keyword":kw,"frequency":f,"sentiment":"negative"})
    pd.DataFrame(kw_rows).to_csv(analysis_dir/"top_keywords_corpus.csv", index=False, encoding="utf-8-sig")

    # Summary JSON
    summary = {
        "total_merchants":          int(len(df)),
        "total_malls":              int(df["mall"].nunique()),
        "total_cities":             int(df.get("city",pd.Series([])).nunique()),
        "avg_rating":               round(float(df["rating"].mean()),2),
        "median_reviews":           round(float(df["reviews"].median()),0),
        "avg_bd_score":             round(float(df["bd_priority_score"].mean()),2),
        "pct_tier1":                round(float((df["tier"]=="Tier 1").mean()*100),1),
        "pct_high_priority":        round(float((df["priority"]=="High").mean()*100),1),
        "pct_positive_sentiment":   round(float((df["sentiment_label"]=="Positive").mean()*100),1),
        "pct_with_phone":           round(float(df["phone"].notna().mean()*100),1) if "phone" in df.columns else 0,
        "merchants_multi_branch":   int((df["branches"]>1).sum()),
        "merchants_delivery_mention":int(df["mentions_delivery"].sum()),
        "whitespace_gaps":          len(gaps),
        "tier_dist":                df["tier"].value_counts().to_dict(),
        "category_dist":            df["category"].value_counts().to_dict(),
        "quadrant_dist":            df["quadrant"].value_counts().to_dict(),
        "top10_bd_merchants":       df.nlargest(10,"bd_priority_score")[["merchant","mall","rating","reviews","bd_priority_score","tier"]].to_dict("records"),
        "top5_malls_by_bd":         mall_agg.nlargest(5,"avg_bd_score")[["mall","merchant_count","avg_rating","avg_bd_score"]].to_dict("records"),
    }
    with open(analysis_dir/"summary_metrics.json","w",encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    logger.info(f"\n{'='*60}")
    logger.info(f"📊 ANALYSIS COMPLETE — {len(df)} merchants")
    logger.info(f"   Tier 1: {summary['pct_tier1']}% | Positive: {summary['pct_positive_sentiment']}%")
    logger.info(f"   Top merchant: {summary['top10_bd_merchants'][0]['merchant']} ({summary['top10_bd_merchants'][0]['bd_priority_score']})")
    logger.info(f"   Files saved to: {analysis_dir}/")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
def main():
    if not GOOGLE_API_KEY:
        print("❌ GOOGLE_PLACES_API_KEY not set in .env!")
        sys.exit(1)

    args = sys.argv[1:]
    fresh        = "--fresh"        in args
    analyze_only = "--analyze-only" in args
    export_only  = "--export-only"  in args
    city_filter  = None
    if "--city" in args:
        idx = args.index("--city")
        if idx+1 < len(args):
            city_filter = args[idx+1]

    target_cities = [c for c in KSA_CITIES if not city_filter or c["name"].lower()==city_filter.lower()]

    print("\n" + "="*65)
    print("   KSA MERCHANT INTELLIGENCE PIPELINE — ALL-IN-ONE")
    print("="*65)
    print(f"   Cities:  {[c['name'] for c in target_cities]}")
    print(f"   API Key: {'✅' if GOOGLE_API_KEY else '❌'}")
    print(f"   Gemini:  {'✅' if GEMINI_API_KEY else '⚠️  skipped'}")
    print("="*65 + "\n")

    excel_output = OUTPUT_DIR / "KSA_Merchants_All.xlsx"

    # ── Analyze only mode ──────────────────────────────────────────────────────
    if analyze_only:
        if excel_output.exists():
            run_analysis(str(excel_output))
        else:
            print(f"❌ No Excel file found at {excel_output}")
        return

    if fresh:
        clear_checkpoint()

    state = load_checkpoint()

    # ── Export only mode ───────────────────────────────────────────────────────
    if export_only:
        if state["enriched"]:
            export_excel(state["enriched"], str(excel_output))
            run_analysis(str(excel_output))
        else:
            print("❌ No enriched data in checkpoint")
        return

    start_time = time.time()

    try:
        for city in target_cities:
            city_name = city["name"]

            if city_name in state.get("cities_done", []):
                print(f"⏭️  {city_name} — already done, skipping")
                continue

            print(f"\n{'─'*65}")
            print(f"🏙️  CITY: {city_name}")
            print(f"{'─'*65}")

            # ── Step 1: Discover malls ─────────────────────────────────────────
            if city_name not in state.get("mall_data", {}):
                print(f"\n🔍 Discovering malls in {city_name}...")
                malls = discover_malls(city)
                state.setdefault("mall_data", {})[city_name] = malls
                save_checkpoint(state)
                print(f"   ✅ Found {len(malls)} malls | {cost_line()}")
            else:
                malls = state["mall_data"][city_name]
                print(f"   ⏩ {len(malls)} malls loaded from checkpoint")

            # ── Step 2: Discover merchants per mall ────────────────────────────
            for mall in malls:
                mall_key = f"{city_name}::{mall['mall_name']}"
                if mall_key in state.get("enriched", {}):
                    n = len(state["enriched"][mall_key])
                    print(f"   ⏩ {mall['mall_name']:<30} {n} merchants (cached)")
                    continue

                print(f"\n   🏪 {mall['mall_name']}")
                merchants = discover_merchants(mall)
                if not merchants:
                    continue
                print(f"      Found {len(merchants)} merchants | {cost_line()}")

                # ── Step 3: Phase 1 enrichment (basic) ────────────────────────
                print(f"      ⚡ Phase 1 (basic details)...")
                enriched_p1 = []
                for i, m in enumerate(merchants):
                    em = enrich_basic(m)
                    enriched_p1.append(em)

                    if (i+1) % SAVE_EVERY == 0:
                        state.setdefault("enriched", {})[mall_key] = enriched_p1[:]
                        save_checkpoint(state)
                        elapsed = time.time() - start_time
                        print(f"      [{i+1}/{len(merchants)}] 💾 saved | {cost_line()} | ⏱ {elapsed/60:.1f}m elapsed")

                # ── Step 4: Phase 2 enrichment (contact) ──────────────────────
                print(f"      📞 Phase 2 (contact details)...")
                enriched_p2 = []
                contact_done = set(state.get("contact_done", []))
                for i, m in enumerate(enriched_p1):
                    pid = m.get("place_id","")
                    if pid and pid not in contact_done:
                        m = enrich_contact(m)
                        if pid: contact_done.add(pid)
                    enriched_p2.append(m)

                    if (i+1) % SAVE_EVERY == 0:
                        state["enriched"][mall_key] = enriched_p2[:]
                        state["contact_done"]        = list(contact_done)
                        save_checkpoint(state)
                        print(f"      [{i+1}/{len(enriched_p1)}] 💾 saved | {cost_line()}")

                state["enriched"][mall_key] = enriched_p2
                state["contact_done"]        = list(contact_done)
                save_checkpoint(state)
                print(f"      ✅ {mall['mall_name']}: {len(enriched_p2)} merchants | {cost_line()}")

            # ── Mark city done ─────────────────────────────────────────────────
            state.setdefault("cities_done",[]).append(city_name)
            save_checkpoint(state)
            print(f"\n✅ {city_name} COMPLETE | {cost_line()}")

    except KeyboardInterrupt:
        print(f"\n\n⏸️  PAUSED at {time.strftime('%H:%M:%S')}")
        print(f"   Progress saved to checkpoint")
        print(f"   {cost_line()}")
        print(f"   Run again to resume from exactly where you stopped.\n")
        sys.exit(0)

    # ── Export Excel ───────────────────────────────────────────────────────────
    print(f"\n📊 Exporting Excel...")
    export_excel(state["enriched"], str(excel_output))

    # ── Run Analysis ───────────────────────────────────────────────────────────
    print(f"\n🧠 Running analysis...")
    run_analysis(str(excel_output))

    # ── Final summary ──────────────────────────────────────────────────────────
    elapsed_total = (time.time() - start_time) / 60
    total_merchants = sum(len(v) for v in state["enriched"].values())

    print(f"\n{'='*65}")
    print(f"✅ ALL DONE!")
    print(f"   Cities:    {state.get('cities_done',[])} ")
    print(f"   Merchants: {total_merchants}")
    print(f"   Time:      {elapsed_total:.1f} minutes")
    print(f"   {cost_line()}")
    print(f"   Excel:     {excel_output}")
    print(f"   Analysis:  {OUTPUT_DIR}/analysis/")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()